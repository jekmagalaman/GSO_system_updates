from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime

from requests.models import ServiceRequest
from .models import WorkAccomplishmentReport


@login_required
def accomplishment_report(request):
    # Completed live requests
    completed_requests = ServiceRequest.objects.filter(status="Completed").order_by("-created_at")

    # Migrated data
    migrated_reports = WorkAccomplishmentReport.objects.all().order_by("-date_started")

    # Normalize into the same format
    reports = []

    # --- Normalize ServiceRequest ---
    for req in completed_requests:
        reports.append({
            "type": "ServiceRequest",
            "requesting_office": req.department,
            "description": req.description,
            "unit": req.unit,
            "date": req.created_at,
            "personnel": [p.get_full_name() for p in req.assigned_personnel.all()] or ["Unassigned"],
            "status": req.status,
            "rating": getattr(req, "rating", None),
        })

    # --- Normalize MigratedReport ---
    for mig in migrated_reports:
        date_value = mig.date_started
        if date_value and not isinstance(date_value, datetime):
            naive_dt = datetime.combine(date_value, datetime.min.time())
            date_value = timezone.make_aware(naive_dt)
        elif isinstance(date_value, datetime) and timezone.is_naive(date_value):
            date_value = timezone.make_aware(date_value)

        # If assigned_personnel is a string, split by comma or wrap in list
        personnel_list = []
        if mig.assigned_personnel:
            if isinstance(mig.assigned_personnel, str):
                personnel_list = [name.strip() for name in mig.assigned_personnel.split(",")]
            else:
                personnel_list = [mig.assigned_personnel]

        reports.append({
            "type": "MigratedReport",
            "requesting_office": mig.requesting_office,
            "description": mig.description,
            "unit": mig.request_type,
            "date": date_value,
            "personnel": personnel_list or ["Unassigned"],
            "status": "Completed",
            "rating": getattr(mig, "rating", None),
        })


    # Apply search filter
    search_query = request.GET.get("user_status")
    if search_query:
        reports = [r for r in reports if search_query.lower() in str(r).lower()]

    # Apply unit filter
    unit_filter = request.GET.get("unit")
    if unit_filter:
        reports = [r for r in reports if r["unit"] == unit_filter]

    # Sort by date (descending)
    reports.sort(key=lambda r: r["date"], reverse=True)

    return render(
        request,
        "gso_office/accomplishment_report/accomplishment_report.html",
        {"reports": reports}
    )














import openpyxl
from django.http import HttpResponse
from django.conf import settings
import os
import calendar

@login_required
def generate_ipmt(request):
    personnel_name = request.GET.get("personnel")
    month_filter = request.GET.get("month")

    if not personnel_name or not month_filter:
        return HttpResponse("Personnel and month are required.", status=400)

    # Parse year-month
    year, month_num = map(int, month_filter.split("-"))
    month_name = f"{calendar.month_name[month_num]} {year}"

    # --- Collect reports (ServiceRequest + MigratedReport) ---
    reports = []  
    # Fill this list with reports as before (your logic here)

    # --- Load the IPMT template ---
    template_path = os.path.join(settings.BASE_DIR, "static", "ipmt_template.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Header info
    ws["B8"] = personnel_name  # Employee Name
    ws["B11"] = month_name     # Month

    # Start writing accomplishments (row 13 downwards)
    start_row = 13
    for i, r in enumerate(reports, start=start_row):
        ws.cell(row=i, column=1).value = r.get("unit")          # Success Indicator
        ws.cell(row=i, column=2).value = r.get("description")   # Actual Accomplishment
        ws.cell(row=i, column=3).value = "Completed"            # Comments / Remarks

    # Return file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"IPMT_{personnel_name}_{month_filter}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
